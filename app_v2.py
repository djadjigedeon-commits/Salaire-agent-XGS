import streamlit as st
import pandas as pd
import openpyxl
import re
import unicodedata
import os
import base64

from supabase import create_client, Client

# ---------- Configuration ----------
st.set_page_config(page_title="Bienvenu ! Votre salaire près de vous", page_icon="💰", layout="centered")

FICHIER_EXCEL = "ETAT_DES_PRIMES_Aout_2026.xlsx"
LOGO_FICHIER = "logo_xgs.png"
FEUILLES_IGNOREES = {"prime manager"}  # comparées en minuscule, sans accent

# ---------- Supabase ----------
@st.cache_resource
def initialiser_supabase():
    """
    Client public : authentification normale.
    Client admin : création des comptes sans envoi d'e-mail.
    La clé secrète reste uniquement côté serveur (Secrets Streamlit).
    """
    try:
        url = st.secrets["SUPABASE_URL"]
        publishable_key = st.secrets["SUPABASE_KEY"]
        secret_key = st.secrets["SUPABASE_SECRET_KEY"]

        if not url or not publishable_key or not secret_key:
            return None, None

        client = create_client(url, publishable_key)
        admin_client = create_client(url, secret_key)

        return client, admin_client

    except Exception:
        return None, None


supabase, supabase_admin = initialiser_supabase()


@st.cache_data
def charger_logo_base64():
    if not os.path.exists(LOGO_FICHIER):
        return None
    with open(LOGO_FICHIER, "rb") as f:
        return base64.b64encode(f.read()).decode()


def appliquer_style():
    logo_b64 = charger_logo_base64()
    fond_logo = f', url("data:image/png;base64,{logo_b64}")' if logo_b64 else ""
    taille_logo = ", 45%" if logo_b64 else ""

    st.markdown(f"""
    <style>
    .stApp {{
        background-image: linear-gradient(135deg, rgba(10, 25, 90, 0.93) 0%, rgba(30, 70, 190, 0.90) 45%, rgba(70, 140, 240, 0.88) 100%){fond_logo};
        background-size: cover{taille_logo};
        background-position: center, center 18%;
        background-repeat: no-repeat, no-repeat;
        background-attachment: fixed, fixed;
    }}
    .block-container {{
        background-color: rgba(255, 255, 255, 0.93);
        border-radius: 18px;
        padding: 2.2rem 2.6rem;
        margin-top: 1.2rem;
        box-shadow: 0 4px 24px rgba(0,0,0,0.25);
    }}
    h1, h2, h3, h4, h5, p, span, label, li,
    .stMarkdown, .stCaption, .stTextInput label, .stSelectbox label {{
        color: #000000 !important;
    }}
    </style>
    """, unsafe_allow_html=True)


appliquer_style()

MOIS_FR = [
    "janvier", "fevrier", "mars", "avril", "mai", "juin",

    "juillet", "aout", "septembre", "octobre", "novembre", "decembre"
]


def sans_accent(texte):
    if not isinstance(texte, str):
        return ""
    nfkd = unicodedata.normalize("NFD", texte)
    return "".join(c for c in nfkd if unicodedata.category(c) != "Mn").lower().strip()


def extraire_mois_annee(nom_feuille):
    """Ex: 'Août 2026' -> ('Aout', 2026, index_mois)."""
    m = re.search(r"(\D+)\s*(\d{4})", nom_feuille)
    if not m:
        return nom_feuille, None, 99
    mot, annee = m.group(1).strip(), int(m.group(2))
    mot_norm = sans_accent(mot)
    idx = MOIS_FR.index(mot_norm) if mot_norm in MOIS_FR else 99
    return mot.strip().capitalize(), annee, idx


def trouver_ligne_entete(ws, max_lignes_scan=15):
    """Cherche la ligne contenant 'Matricule' ou 'Nom et prénoms'."""
    for r in range(1, max_lignes_scan + 1):
        valeurs = [sans_accent(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        if "matricule" in valeurs or any("nom et prenom" in v for v in valeurs):
            return r
    return None


def mapper_colonnes(ws, ligne_entete, mois_norm):
    """Associe chaque champ standard à un numéro de colonne, via correspondance souple sur l'intitulé."""
    mapping = {}
    for c in range(1, ws.max_column + 1):
        intitule = sans_accent(ws.cell(row=ligne_entete, column=c).value)
        if not intitule:
            continue
        if "matricule" in intitule:
            mapping["Matricule"] = c
        elif "mot de passe" in intitule or "password" in intitule:
            mapping["MotDePasse"] = c
        elif "nom et prenom" in intitule:
            mapping["Nom"] = c
        elif intitule == "projet" or "projet" in intitule:
            mapping.setdefault("Projet", c)
        elif "superviseur" in intitule:
            mapping["Superviseur"] = c
        elif intitule == "poste":
            mapping["Poste"] = c
        elif "salaire de base" in intitule:
            mapping["Salaire_Base"] = c
        elif "net a payer" in intitule and "total" not in intitule:
            mapping["Net_A_Payer"] = c
        elif "coach" in intitule:
            mapping["Prime_Coach"] = c
        elif "prime" in intitule:
            mapping["Prime_Anterieure"] = c
        elif "total" in intitule and mois_norm and mois_norm in intitule:
            mapping["Total_Mois"] = c
        elif "total" in intitule and "sans formule" in intitule:
            mapping.setdefault("Total_Mois_Precedent", c)
    return mapping


@st.cache_data(ttl=60)
def charger_donnees():
    if not os.path.exists(FICHIER_EXCEL):
        return None, []

    wb_valeurs = openpyxl.load_workbook(FICHIER_EXCEL, data_only=True)
    lignes = []
    feuilles_lues = []

    for nom_feuille in wb_valeurs.sheetnames:
        if sans_accent(nom_feuille) in FEUILLES_IGNOREES:
            continue

        mois_label, annee, ordre_mois = extraire_mois_annee(nom_feuille)
        if annee is None:
            continue  # feuille qui ne ressemble pas à un mois -> ignorée

        ws = wb_valeurs[nom_feuille]
        ligne_entete = trouver_ligne_entete(ws)
        if ligne_entete is None:
            continue

        mapping = mapper_colonnes(ws, ligne_entete, sans_accent(mois_label))
        if "Nom" not in mapping:
            continue

        feuilles_lues.append(nom_feuille)

        for r in range(ligne_entete + 1, ws.max_row + 1):
            nom = ws.cell(row=r, column=mapping["Nom"]).value
            if not nom or not str(nom).strip():
                continue  # ligne vide ou ligne de total en bas de tableau

            def val(champ):
                col = mapping.get(champ)
                return ws.cell(row=r, column=col).value if col else None

            matricule = val("Matricule")
            lignes.append({
                "Matricule": str(matricule).strip() if matricule not in (None, "") else None,
                "Nom": str(nom).strip(),
                "Mois": mois_label,
                "Annee": annee,
                "ordre_mois": ordre_mois,
                "Projet": val("Projet"),
                "Poste": val("Poste"),
                "Salaire_Base": val("Salaire_Base"),
                "Net_A_Payer": val("Net_A_Payer"),
                "Prime_Anterieure": val("Prime_Anterieure"),
                "Prime_Coach": val("Prime_Coach"),
                "Total_Mois": val("Total_Mois"),
            })

    if not lignes:
        return None, feuilles_lues

    df = pd.DataFrame(lignes)
    return df, feuilles_lues


def email_auth_depuis_matricule(matricule):
    """Identifiant technique utilisé par Supabase Auth à partir du matricule."""
    matricule_normalise = str(matricule).strip().lower()
    return f"{matricule_normalise}@xgs.local"


def creer_compte_supabase(matricule, mot_de_passe, nom=None):
    """
    Crée un compte Supabase avec l'API Admin.
    Aucun e-mail n'est envoyé.
    email_confirm=True considère immédiatement l'identifiant comme confirmé.
    """
    email = email_auth_depuis_matricule(matricule)

    return supabase_admin.auth.admin.create_user({
        "email": email,
        "password": str(mot_de_passe),
        "email_confirm": True,
        "user_metadata": {
            "matricule": str(matricule).strip(),
            "nom": str(nom or "").strip(),
        },
    })


def authentifier_ou_creer_compte(matricule, mot_de_passe, nom=None):
    """
    Première connexion :
      - crée le compte avec l'API Admin ;
      - aucun e-mail n'est envoyé ;
      - email_confirm=True ;
      - connexion immédiate.

    Connexions suivantes :
      - vérifie le mot de passe avec Supabase Auth.
    """
    if supabase is None or supabase_admin is None:
        return False, (
            "Supabase n'est pas configuré. Vérifiez les secrets de l'application "
            "(SUPABASE_URL, SUPABASE_KEY et SUPABASE_SECRET_KEY)."
        )

    matricule = str(matricule).strip()
    mot_de_passe = str(mot_de_passe)

    if not matricule:
        return False, "Veuillez saisir votre matricule."

    if not mot_de_passe:
        return False, "Veuillez saisir votre mot de passe."

    email = email_auth_depuis_matricule(matricule)

    # --------------------------------------------------------
    # 1. Première connexion : tentative de création
    # --------------------------------------------------------
    try:
        creer_compte_supabase(
            matricule=matricule,
            mot_de_passe=mot_de_passe,
            nom=nom,
        )

        # Le compte vient d'être créé.
        # email_confirm=True => aucune confirmation par e-mail.
        supabase.auth.sign_in_with_password({
            "email": email,
            "password": mot_de_passe,
        })

        return True, "Compte créé et connexion réussie."

    except Exception as e:
        message = str(e).lower()

        # Si le compte existe déjà, ce n'est pas une erreur fonctionnelle :
        # on passe à l'authentification normale.
        compte_existant = (
            "already registered" in message
            or "already exists" in message
            or "user already exists" in message
            or "duplicate" in message
            or "email address is already registered" in message
        )

        if not compte_existant:
            return False, f"Impossible de créer le compte : {e}"

    # --------------------------------------------------------
    # 2. Compte existant : connexion normale
    # --------------------------------------------------------
    try:
        connexion = supabase.auth.sign_in_with_password({
            "email": email,
            "password": mot_de_passe,
        })

        if getattr(connexion, "session", None) is not None:
            return True, "Connexion réussie."

        return False, "Mot de passe incorrect."

    except Exception:
        return False, "Mot de passe incorrect."

def formater_fcfa(valeur):
    try:
        return f"{valeur:,.0f} FCFA".replace(",", " ")
    except (ValueError, TypeError):
        return "-" if valeur in (None, "") else valeur


# ---------- Interface ----------
st.title("Bienvenu ! Votre salaire près de vous")
st.caption(
    "Entrez votre matricule et votre mot de passe pour consulter le détail, mois par mois. "
    "Lors de votre première connexion, votre compte est créé automatiquement."
)

df, feuilles_lues = charger_donnees()

if df is None:
    st.error(
        f"Impossible de trouver des données exploitables dans '{FICHIER_EXCEL}'. "
        "Vérifiez que le fichier est présent et que la colonne Matricule est bien remplie."
    )
    st.stop()

with st.form("connexion"):
    matricule_saisi = st.text_input("Votre matricule", placeholder="Ex : A001").strip()
    mot_de_passe_saisi = st.text_input("Votre mot de passe", type="password").strip()
    valider = st.form_submit_button("Voir mon salaire")

if valider:
    st.session_state["connecte"] = False

    if not matricule_saisi or not mot_de_passe_saisi:
        st.warning("Veuillez renseigner votre matricule et votre mot de passe.")
    else:
        correspondances = df[
            df["Matricule"].astype(str).str.strip().str.lower()
            == matricule_saisi.strip().lower()
        ]

        if correspondances.empty:
            st.error("Matricule inconnu. Contactez le service RH si besoin.")
        else:
            nom_utilisateur = ""
            if "Nom" in correspondances.columns:
                valeur_nom = correspondances.iloc[0]["Nom"]
                nom_utilisateur = "" if pd.isna(valeur_nom) else str(valeur_nom).strip()

            authentifie, message = authentifier_ou_creer_compte(
                matricule_saisi,
                mot_de_passe_saisi,
                nom_utilisateur
            )

            if authentifie:
                st.session_state["connecte"] = True
                st.session_state["matricule_connecte"] = matricule_saisi.strip()
                st.success(message)
            else:
                st.error(message)

if st.session_state.get("connecte"):
    resultats = df[df["Matricule"] == st.session_state["matricule_connecte"]].copy()

    if resultats.empty:
        st.warning(
            "Aucune donnée trouvée pour ce matricule. "
            "Vérifiez la saisie, ou contactez le service RH si votre matricule n'a pas encore été renseigné."
        )
    else:
        nom_complet = resultats.iloc[0]["Nom"]
        st.success(f"Bienvenue, **{nom_complet}** (Matricule : {st.session_state['matricule_connecte']})")

        resultats = resultats.sort_values(["Annee", "ordre_mois"])

        annees_disponibles = sorted(resultats["Annee"].dropna().unique().tolist(), reverse=True)
        annee_choisie = st.selectbox("Année", annees_disponibles) if len(annees_disponibles) > 1 else annees_disponibles[0]

        vue_annee = resultats[resultats["Annee"] == annee_choisie]

        colonnes_affichees = ["Mois", "Salaire_Base", "Net_A_Payer", "Prime_Anterieure", "Prime_Coach", "Total_Mois"]
        noms_lisibles = {
            "Mois": "Mois",
            "Salaire_Base": "Salaire de base",
            "Net_A_Payer": "Net à payer (hors primes)",
            "Prime_Anterieure": "Prime (mois précédent)",
            "Prime_Coach": "Prime coach métier",
            "Total_Mois": "Total du mois",
        }
        tableau = vue_annee[colonnes_affichees].rename(columns=noms_lisibles).reset_index(drop=True)

        for col in ["Salaire de base", "Net à payer (hors primes)", "Prime (mois précédent)", "Prime coach métier", "Total du mois"]:
            tableau[col] = tableau[col].apply(formater_fcfa)

        st.subheader(f"Détail par mois — {annee_choisie}")
        st.dataframe(tableau, hide_index=True, use_container_width=True)

        mois_disponibles = vue_annee["Mois"].tolist()
        mois_choisi = st.selectbox("Voir le détail d'un mois précis", mois_disponibles)
        ligne = vue_annee[vue_annee["Mois"] == mois_choisi].iloc[0]

        c1, c2, c3 = st.columns(3)
        c1.metric("Salaire de base", formater_fcfa(ligne["Salaire_Base"]))
        c2.metric("Net à payer", formater_fcfa(ligne["Net_A_Payer"]))
        c3.metric("Total du mois", formater_fcfa(ligne["Total_Mois"]))

        if ligne.get("Projet") or ligne.get("Poste"):
            st.caption(f"Projet : {ligne.get('Projet', '-')}  |  Poste : {ligne.get('Poste', '-')}")

        if st.button("Se déconnecter"):
            st.session_state["connecte"] = False
            st.session_state.pop("matricule_connecte", None)
            st.rerun()

st.divider()
st.caption("Cette page est en lecture seule : les données de salaire sont consultées depuis le fichier de référence. L’authentification est gérée par Supabase.")
